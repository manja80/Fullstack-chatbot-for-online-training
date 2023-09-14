import openpyxl
from openpyxl import load_workbook
import re
import psycopg2
# Function to validate email format
def is_valid_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email)

# Function to save data to the PostgreSQL database securely
def save_to_database(name, email, mobile):
    db_config = {
        'dbname': 'postgres',
        'user': 'postgres',
        'password': 'manoj1234',
        'host': 'localhost',  # Use 'localhost' if the database is on your local machine
        'port': '5433',  # Default PostgreSQL port is 5432
    }

    try:
        conn = psycopg2.connect(**db_config)
        cursor = conn.cursor()

        # Insert data into the database
        insert_query = "INSERT INTO training_details (name, email, mobile) VALUES (%s, %s, %s)"
        cursor.execute(insert_query, (name, email, mobile))
        conn.commit()

        cursor.close()
        conn.close()
    except Exception as e:
        print(f"An error occurred while saving data to the database: {str(e)}")


# Function to initialize or load the Excel workbook
def init_or_load_excel_workbook():
    try:
        workbook = load_workbook("training_details.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Training Details"
        worksheet.append(["Name", "Email", "Mobile"])
        workbook.save("training_details.xlsx")
    return workbook

# Function to save data to the Excel worksheet securely
def save_to_excel_safely(workbook, name, email, mobile):
    worksheet = workbook.active
    worksheet.append([name, email, mobile])
    workbook.save("training_details.xlsx")

# Main chatbot function
def chatbot():
    print("Welcome to the Web Development Training Chatbot!")

    while True:
        response = input("Are you interested in web development training? (yes/no): ").strip().lower()

        if response == "yes":
            name = input("Great! Please provide your name: ").strip()
            email = input("Please provide your email address: ").strip()

            if is_valid_email(email):
                mobile = input("Please provide your mobile number: ").strip()

                try:
                    # Initialize or load the Excel workbook
                    workbook = init_or_load_excel_workbook()

                    # Save data to Excel securely
                    save_to_excel_safely(workbook, name, email, mobile)

                    print("Thank you for providing your details. We will contact you soon.")
                    break
                except Exception as e:
                    print(f"An error occurred: {str(e)}")
                    print("Please try again later.")
                    break
            else:
                print("Invalid email format. Please provide a valid email address.")
        elif response == "no":
            print("Okay, if you change your mind, feel free to come back.")
            break
        else:
            print("Invalid response. Please enter 'yes' or 'no'.")

if __name__ == "__main__":
    chatbot()
